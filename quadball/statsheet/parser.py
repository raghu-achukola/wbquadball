from io import BytesIO
from quadball.schema.statsheet.statsheet_pb2 import StatSheetPossession
from google.protobuf.json_format import ParseDict
# Typing
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from typing import Generator

DEFAULT_STATSHEET_START = (7,1)

def parse_file(file_location:str) -> list:
    """
        High level method to convert a file location into a stats json 
    """
    #TODO Inc 2: fill method with functionality
    return []

def open_file(file_location:str, location_type = None) -> BytesIO:
    """
        The output of both a BytesIO object and a S3 StreamingBody
        can be accessed by .read()

        In this increment we will keep the default functionality the
        local one
    """
    #TODO: add s3 functionality
    with open(file_location,'rb') as f: 
        return BytesIO(f.read())

def get_cell_group_values(ws:Worksheet ,starter_row,starter_col) -> list:
    """
        Given a worksheet and a starting row and column. this function 
        will output the values left -> right then up -> down of the 2x3 rectangle
        starting at the starter cell. This if placed correctly represents a 
        'block' in the statsheet representing a possesssion. 
        The default on the statsheet should be A7 aka (7,1)
    """
    return [
        ws.cell(starter_row + r, starter_col +c).value 
        for r in range(2)
        for c in range(3)
    ]

def gen_statsheet_values(ws:Worksheet, starter_cell:Cell) -> Generator[list,None,None]:
    """
        Generator method
        Bounces from possession block to possession block yielding the 
        statsheet values in the 2x3 possession block each time
    """ 
    first = True
    while True: 
        starter_cell = starter_cell.offset(0,0 if first else 3)
        first = False
        # NOTE: If we ever replace the A or B, this will BREAK
        if starter_cell.offset(0,1).value not in ('A','B'): 
            # We ran out of the columns in this row
            # Start over in the next row 
            # NOTE: we are assuming the cell starts over at 1 (no padding)
            starter_cell = ws.cell(starter_cell.row +3, 1)
            if starter_cell.offset(0,1).value not in ('A','B'): 
                # We ran out of rows
                break
        yield get_cell_group_values( ws, starter_cell.row, starter_cell.column)


def gen_statsheet_possessions(ws:Worksheet, starter_cell:Cell = None) -> Generator[StatSheetPossession, None, None]:
    """
        Generator method (for now)
        Bounces from posession to possession and converts to our Schema
        StatsheetPossession object
    """
    previous_offense = None
    first_possession = True
    if starter_cell is None: 
        starter_cell = ws.cell(*DEFAULT_STATSHEET_START) #default starting cell 
    for extras, offense, end_time, result, primary, secondary in gen_statsheet_values(ws, starter_cell):
        assert offense != previous_offense
        if not (result or first_possession): 
            # Game over (i mean not really but) 
            break
        elif not result: 
            pass
        else: 
            first_possession=False
            previous_offense=offense
            yield ParseDict(
                    {
                        'extras':str(extras) if extras is not None else '',
                        'offense':str(offense) if offense is not None else '',
                        'end_time':str(end_time) if end_time is not None else '',
                        'result':str(result) if result is not None else '',
                        'primary':str(primary) if primary is not None else '',
                        'secondary':str(secondary) if secondary is not None else '' 
                    },
                    StatSheetPossession()
                )

