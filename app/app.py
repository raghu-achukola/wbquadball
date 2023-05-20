# Import flask functions
from flask import Flask, render_template,request
import requests
import pandas as pd
import json
from pkgutil import get_data
import openpyxl
from io import BytesIO
app = Flask(__name__)

PUBLIC_METADATA_BUCKET_URL = 'https://wbquadball-uw2-public-metadata-storage.s3.us-west-2.amazonaws.com' 
# We want to store the html somewhere else though, not just as a giant string. 
# Enter render_template
@app.route('/')
def root(): 
    return render_template('index.html') 

@app.route('/all')
def all(): 
    LEAGUE_PREFIX = PUBLIC_METADATA_BUCKET_URL + '/db_snapshot/{}.json'
    result = {collection_name: requests.get(LEAGUE_PREFIX.format(collection_name)).json() for collection_name in ('leagues','seasons','tournaments','teams','games')}
    games_x_teams = []
    team_dict = {team['_id']:team['team_id'] for team in result['teams']}
    for g in result['games']: 
        game = g.copy()
        game['team_a_name'] = team_dict[game['winning_team_id']]
        game['team_b_name'] = team_dict[game['losing_team_id']]
        games_x_teams.append(game)
    result['games'] = games_x_teams
    return result

@app.route('/statsheet')
def gen_statsheet():
    print('AAAH')
    print(request.args)
    args = request.args
    ## Theres a better way to be doing this i can't remember it 
    # off the top of my head    
    excel_bytes = get_data('quadball','resources/statsheet/STATSHEET_TEMPLATE.xlsx')
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    metadata = wb['METADATA'] 
    metadata.cell(1,2).value = args.get('season_id')
    metadata.cell(2,2).value = args.get('game_id')
    metadata.cell(3,2).value = args.get('tournament_id')
    metadata.cell(4,2).value = args.get('team_a_id')
    metadata.cell(5,2).value = args.get('team_b_id')
    possessions = wb['POSSESSIONS']
    possessions.cell(2,3).value = args.get('team_a_name')
    possessions.cell(4,3).value = args.get('team_b_name')
    # Transform the workbook
    buffer = BytesIO()
    buffer.seek(0)
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
# If name is main, run flask 
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug = True)